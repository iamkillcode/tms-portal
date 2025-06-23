# WARNING: This file is maintained for backward compatibility only.
# All views have been moved to the views/ directory for better organization.
# Please use the modular imports from the views/ directory instead.

# Re-export all views from the views/ directory
from .views import *

from .models import (
    TenderTracker, Department, Category, Tender, TenderItem,
    VendorBid, FrameworkAgreement, Vendor, ISONumber, ISOTracker,
    Division, UserProfile, BreakfastItem, Order, OrderItem,
    Chemical, ChemicalSpecification, Task, TaskCategory, TaskComment
)

from .forms import (
    CustomUserCreationForm, TenderItemForm, VendorBidForm,
    FrameworkAgreementForm, ChemicalForm, ChemicalSpecificationForm,
    ChemicalImportForm, TaskForm, TaskCategoryForm, TaskCommentForm,
    UserProfileForm, VendorForm
)
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd

# For type annotations and imports needed in other files
from django.shortcuts import render, redirect, get_object_or_404
from datetime import datetime, timezone, timedelta
from django.http import JsonResponse, HttpResponse
from django.contrib.auth import login, authenticate
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.forms import AuthenticationForm, UserCreationForm
from django.contrib import messages
from django.contrib.auth.models import User
from django import forms
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.db.models.functions import ExtractMonth
from django.urls import reverse
from django.utils import timezone as django_timezone
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView

from .models import (
    TenderTracker, Department, Category, Tender, TenderItem,
    VendorBid, FrameworkAgreement, Vendor, ISONumber, ISOTracker,
    Division, UserProfile, BreakfastItem, Order, OrderItem,
    Chemical, ChemicalSpecification, Task, TaskCategory, TaskComment
)

from .forms import (
    CustomUserCreationForm, TenderItemForm, VendorBidForm,
    FrameworkAgreementForm, ChemicalForm, ChemicalSpecificationForm,
    ChemicalImportForm, TaskForm, TaskCategoryForm, TaskCommentForm,
    UserProfileForm, VendorForm
)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd

def has_admin_role(user):
    if user.is_superuser:
        return True
    allowed_groups = {'Admin', 'Head of Department', 'Head of Unit'}
    user_groups = set(user.groups.values_list('name', flat=True))
    return bool(allowed_groups & user_groups)

def has_user_role(user):
    if has_admin_role(user):
        return True
    allowed_groups = {'Team Lead', 'Officer'}
    user_groups = set(user.groups.values_list('name', flat=True))
    return bool(allowed_groups & user_groups)

# Example usage for an admin-only view:
@user_passes_test(has_admin_role)
def admin_dashboard_view(request):
    # ...existing code for admin dashboard...
    pass

# Example usage for a user-level view:
@user_passes_test(has_user_role)
def user_dashboard_view(request):
    # ...existing code for user dashboard...
    pass

# def create_missing_profiles():
#     for user in User.objects.all():
#         if not hasattr(user, 'profile'):
#             UserProfile.objects.create(
#                 user=user,
#                 full_name=f"{user.first_name} {user.last_name}".strip() or user.username
#             )

# if __name__ == '__main__':
#     create_missing_profiles()

# Tender number generator view
@login_required
@user_passes_test(has_user_role)
def tender_generator_view(request):
    # Ensure user has a profile
    if not hasattr(request.user, 'profile'):
        UserProfile.objects.create(
            user=request.user,
            full_name=f"{request.user.first_name} {request.user.last_name}".strip() or request.user.username
        )
    
    tender_number = None
    error = None
    tender_description = None  # Variable to hold the description

    if request.method == "POST":
        try:
            # Get form inputs
            department_code = request.POST.get("department_code")
            category_code = request.POST.get("category_code")
            procurement_type = request.POST.get("procurement_type")
            tender_description = request.POST.get("tender_description")  # Get the description field
            lot_number = request.POST.get("lot_number")
            amendment_number = request.POST.get("amendment_number")
            call_off_number = request.POST.get("call_off_number")

            # Check if inputs are missing or invalid
            if not department_code or not category_code or not procurement_type:
                raise ValueError("All fields must be provided.")

            prefix = "FDA"
            year = datetime.now().year

            # Retrieve or create a tracker for the current year
            tracker, created = TenderTracker.objects.get_or_create(
                year=year,
                defaults={'last_sequence': 0}  # Initialize sequence if new
            )

            # Increment the global sequence
            tracker.last_sequence += 1
            tracker.save()

            # Generate the tender number
            sequential_number = f"{tracker.last_sequence:04}"  # Zero-padded to 4 digits
            lot_suffix = f" ({int(lot_number):02d})" if lot_number else ""
            amendment_suffix = f" (A{amendment_number})" if amendment_number else ""
            call_off_suffix = f" (CO{call_off_number})" if call_off_number else ""
            tender_number = f"{prefix}/{department_code}/{year}/{category_code}/{procurement_type}-{sequential_number}{lot_suffix}{amendment_suffix}{call_off_suffix}"

            # Save Tender into the database
            department = Department.objects.get(code=department_code)
            category = Category.objects.get(code=categoryCode)

            tender = Tender.objects.create(
                tender_number=tender_number,
                description=tender_description,
                department=department,
                category=category,
                user=request.user  # Automatically assign logged-in user
            )
            messages.success(request, f'Tender {tender_number} created successfully!')

        except Department.DoesNotExist:
            messages.error(request, 'Invalid department code selected.')
            error = "Invalid department code"
        except Category.DoesNotExist:
            messages.error(request, 'Invalid category code selected.')
            error = "Invalid category code"
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
            error = str(e)

    # Fetch department codes for the dropdown
    department_codes = Department.objects.values_list('code', flat=True)
    categories = Category.objects.all()
    category_data = [{"id": category.code, "text": category.name} for category in categories]

    # Fetch tenders, ordered by most recent
    tenders = Tender.objects.all().order_by('-created_at')

    # Add pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(tenders, 10)  # Show 10 tenders per page
    try:
        tenders = paginator.page(page)
    except:
        tenders = paginator.page(1)

    return render(request, "tender_generator.html", {
        "tender_number": tender_number,
        "error": error,
        "department_codes": department_codes,
        "tender_description": tender_description,  # Pass the description to the template
        "category_data": category_data,
        "tenders": tenders,  # Pass tenders to the template
        "is_paginated": True if tenders.has_other_pages() else False,
    })

class CustomUserCreationForm(UserCreationForm):
    email = forms.EmailField(required=True)

    class Meta:
        model = User
        fields = ("username", "email", "password1", "password2")

# Register view
def register_view(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(request, 'Account created successfully!')
            return redirect('login')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = CustomUserCreationForm()
    return render(request, 'register.html', {'form': form})

# Login view
def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
        
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                next_url = request.GET.get('next', reverse('dashboard'))
                return redirect(next_url)
            else:
                messages.error(request, 'Invalid username or password.')
    else:
        form = AuthenticationForm()
    
    return render(request, 'login.html', {'form': form})

# Home view (after login)
def home_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    return redirect('login')

@login_required
@user_passes_test(has_user_role)
def tender_activity_view(request):
    search_query = request.GET.get('search', '')
    
    tenders = Tender.objects.all()
    if search_query:
        tenders = tenders.filter(
            Q(tender_number__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(department__name__icontains=search_query)
        )
    
    tenders = tenders.order_by('-created_at')
    
    # Add pagination here too
    paginator = Paginator(tenders, 10)
    page = request.GET.get('page')
    tenders = paginator.get_page(page)
    
    return render(request, 'tender_activity.html', {
        "tenders": tenders,
        "search_query": search_query
    })

@login_required
@user_passes_test(has_user_role)
def tender_list_view(request):
    search_query = request.GET.get('search', '')
    
    tenders = Tender.objects.all()
    if search_query:
        tenders = tenders.filter(
            Q(tender_number__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(department__name__icontains=search_query)
        )
    
    tenders = tenders.order_by('-created_at')
    
    # Get all vendors for the dropdown
    vendors = Vendor.objects.all().order_by('name')
    
    paginator = Paginator(tenders, 10)
    page = request.GET.get('page')
    tenders = paginator.get_page(page)
    
    return render(request, 'tender_list.html', {
        'tenders': tenders,
        'vendors': vendors,
        'search_query': search_query
    })

@login_required
@user_passes_test(has_user_role)
def tender_update_view(request, tender_id):
    tender = get_object_or_404(Tender, id=tender_id)
    
    if request.method == 'POST':
        # Update tender with form data
        tender.category = request.POST.get('category')
        tender.status = request.POST.get('status')
        tender.invitation_date = request.POST.get('invitation_date') or None
        tender.closing_date = request.POST.get('closing_date') or None
        tender.evaluation_date = request.POST.get('evaluation_date') or None
        tender.contract_date = request.POST.get('contract_date') or None
        tender.currency = request.POST.get('currency')
        tender.contract_amount = request.POST.get('contract_amount') or None
        
        # Keep vendor_name field for backward compatibility
        vendor_name = request.POST.get('vendor_name')
        tender.vendor_name = vendor_name
        
        # Create relationship with vendor using proper ForeignKey
        vendor_id = request.POST.get('vendor')
        if vendor_id:
            tender.vendor = get_object_or_404(Vendor, id=vendor_id)
        else:
            tender.vendor = None
        
        tender.save()
        messages.success(request, 'Tender details updated successfully.')
        
    return redirect('tender-list')

@login_required
@user_passes_test(has_admin_role)
def export_tenders_view(request):
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tender List"

    # Define headers and styles
    headers = [
        'Tender Number',
        'Description',
        'Department',
        'Category',
        'Status',
        'Officer',
        'Invitation Date',
        'Closing Date',
        'Evaluation Date',
        'Contract Date',
        'Currency',
        'Contract Amount',
        'Vendor/Consultant',
        'PO Number',
        'PO Date',
        'SRA Number',
        'SRA Date',
        'Payment Amount',
        'File Name',
        'File Number',
        'Created Date'
    ]

    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Get all tenders
    tenders = Tender.objects.all().order_by('-created_at')

    # Write data with timezone handling
    for row, tender in enumerate(tenders, 2):
        ws.cell(row=row, column=1, value=tender.tender_number)
        ws.cell(row=row, column=2, value=tender.description)
        ws.cell(row=row, column=3, value=tender.department.name)
        ws.cell(row=row, column=4, value=tender.category)
        ws.cell(row=row, column=5, value=tender.status)
        
        try:
            officer_name = tender.user.profile.full_name
        except:
            officer_name = tender.user.get_full_name() or tender.user.username
        
        ws.cell(row=row, column=6, value=officer_name)
        
        # Handle dates with timezone conversion
        date_fields = [
            tender.invitation_date,
            tender.closing_date,
            tender.evaluation_date,
            tender.contract_date,
            tender.po_date,
            tender.sra_date,
            tender.payment_memo_date
        ]
        
        for col, date in enumerate(date_fields, 7):
            if date:
                # Date fields don't need timezone handling as they're already timezone-naive
                ws.cell(row=row, column=col, value=date)

        # Handle numeric and text fields
        ws.cell(row=row, column=11, value=tender.currency)
        ws.cell(row=row, column=12, value=tender.contract_amount)
        ws.cell(row=row, column=13, value=tender.vendor_name)
        ws.cell(row=row, column=14, value=tender.po_number)
        ws.cell(row=row, column=16, value=tender.sra_number)
        ws.cell(row=row, column=18, value=tender.payment_amount)
        ws.cell(row=row, column=19, value=tender.file_name)
        ws.cell(row=row, column=20, value=tender.file_number)

        # Handle created_at datetime with timezone conversion
        if tender.created_at.tzinfo:
            created_at = tender.created_at.astimezone(timezone.utc).replace(tzinfo=None)
        else:
            created_at = tender.created_at

        ws.cell(row=row, column=21, value=created_at)

        # Format the date cell
        date_cell = ws.cell(row=row, column=21)
        date_cell.number_format = 'YYYY-MM-DD HH:MM:SS'

    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=tender_list_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    try:
        wb.save(response)
    except Exception as e:
        messages.error(request, f"Error exporting tenders: {str(e)}")
        return redirect('tender-list')

    return response

@login_required
@user_passes_test(has_user_role)
def shop_view(request):
    breakfast_items = BreakfastItem.objects.filter(available=True)
    return render(request, 'shop.html', {'breakfast_items': breakfast_items})

@login_required
@user_passes_test(has_user_role)
def order_list_view(request):
    orders = Order.objects.filter(user=request.user).order_by('-created_at')
    return render(request, 'order_list.html', {'orders': orders})

@login_required
@user_passes_test(has_user_role)
def add_to_order_view(request, item_id):
    if request.method == 'POST':
        item = get_object_or_404(BreakfastItem, id=item_id)
        quantity = int(request.POST.get('quantity', 1))
        
        # Get or create pending order
        order, created = Order.objects.get_or_create(
            user=request.user,
            status='pending'
        )
        
        # Add item to order
        OrderItem.objects.create(
            order=order,
            item=item,
            quantity=quantity,
            price=item.price
        )
        
        # Update total amount
        order.total_amount = sum(
            item.price * item.quantity 
            for item in order.orderitem_set.all()
        )
        order.save()
        
        messages.success(request, f'{quantity}x {item.name} added to your order!')
        
    return redirect('shop')

@login_required
@user_passes_test(has_user_role)
def iso_generator_view(request, tender_id=None):
    if tender_id:
        tender = get_object_or_404(Tender, id=tender_id)
    else:
        tender = None

    if request.method == "POST":
        try:
            division_code = request.POST.get("division_code")
            department_code = request.POST.get("department_code")
            letter_type = request.POST.get("letter_type")
            tender_id = request.POST.get("tender_id")

            if not all([division_code, department_code, letter_type, tender_id]):
                raise ValueError("All fields must be provided.")

            tender = get_object_or_404(Tender, id=tender_id)
            division = get_object_or_404(Division, code=division_code)
            department = get_object_or_404(Department, code=department_code)

            # Generate ISO number
            prefix = "FDA"
            year = datetime.now().year
            year_short = str(year)[-2:]
            
            # Get next sequence number
            sequence = ISOTracker.get_next_sequence(year)
            sequence_str = str(sequence).zfill(4)

            iso_number = f"{prefix}/{division.code}/{department.code}/{letter_type}/{year_short}/{sequence_str}"

            # Create ISO number record
            iso = ISONumber.objects.create(
                iso_number=iso_number,
                officer=request.user,
                tender=tender,
                division=division,
                department=department,
                letter_type=letter_type,
                description=tender.description
            )

            messages.success(request, f'ISO Number generated successfully: {iso_number}')
            return redirect('iso-detail', iso_id=iso.id)

        except Exception as e:
            messages.error(request, str(e))

    divisions = Division.objects.all()
    if not divisions.exists():
        messages.warning(request, "No divisions found. Please add divisions to the database.")
    
    context = {
        'divisions': divisions,
        'departments': Department.objects.all(),
        'tender': tender,
        'tenders': Tender.objects.all() if not tender else None
    }
    
    return render(request, 'iso_generator.html', context)

@login_required
@user_passes_test(has_user_role)
def iso_detail_view(request, iso_id):
    iso = get_object_or_404(ISONumber, id=iso_id)
    return render(request, 'iso_detail.html', {'iso': iso})

@login_required
@user_passes_test(has_user_role)
def iso_list_view(request):
    from tender_app.htmx_utils import htmx_template
    
    @htmx_template('iso_list.html', 'htmx/iso_list_rows.html')
    def _iso_list(request):
        isos = ISONumber.objects.all().order_by('-date_created')
        search_query = request.GET.get('search', '')
        
        if search_query:
            isos = isos.filter(
                Q(iso_number__icontains=search_query) |
                Q(description__icontains=search_query) |
                Q(division__name__icontains=search_query) |
                Q(department__name__icontains=search_query)
            )
        
        paginator = Paginator(isos, 10)
        page = request.GET.get('page')
        isos_page = paginator.get_page(page)
        return {
            'isos': isos_page,
            'search_query': search_query
        }
    
    return _iso_list(request)

@login_required
@user_passes_test(has_user_role)
def dashboard_view(request):    # Get counts
    total_tenders = Tender.objects.count()
    total_isos = ISONumber.objects.count()
    total_vendors = Vendor.objects.count()
    
    # Get recent items (last 30 days)
    thirty_days_ago = timezone.now() - timedelta(days=30)
    recent_tenders = Tender.objects.filter(
        created_at__gte=thirty_days_ago
    ).order_by('-created_at')[:5]
    
    recent_isos = ISONumber.objects.filter(
        date_created__gte=thirty_days_ago
    ).order_by('-date_created')[:5]
    
    # Get tender statistics by department
    department_stats = Tender.objects.values('department__name')\
        .annotate(count=Count('id'))\
        .order_by('-count')[:5]
    
    context = {
        'total_tenders': total_tenders,
        'total_isos': total_isos,
        'recent_tenders': recent_tenders,
        'recent_isos': recent_isos,
        'department_stats': department_stats,
    }
    
    return render(request, 'dashboard.html', context)

class TenderDetailView(DetailView):
    model = Tender
    template_name = 'tender_app/tender_detail.html'
    context_object_name = 'tender'

@login_required
@user_passes_test(has_user_role)
def search_view(request):
    search_query = request.GET.get('search', '')
    
    tenders = []
    isos = []
    
    if search_query:
        tenders = Tender.objects.filter(
            Q(tender_number__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(department__name__icontains=search_query) |
            Q(category__icontains=search_query)
        ).order_by('-created_at')
        
        isos = ISONumber.objects.filter(
            Q(iso_number__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(division__name__icontains=search_query) |
            Q(department__name__icontains=search_query)
        ).order_by('-date_created')
    
    return render(request, 'search.html', {
        'search_query': search_query,
        'tenders': tenders,
        'isos': isos
    })

@login_required
@user_passes_test(has_admin_role)
def reports_view(request):
    # Get overall statistics
    total_tenders = Tender.objects.count()
    total_isos = ISONumber.objects.count()
    completed_tenders = Tender.objects.filter(status='Completed').count()
    
    # Get tender statistics by department
    department_stats = Tender.objects.values('department__name')\
        .annotate(count=Count('id'))\
        .order_by('-count')
    
    # Get tender statistics by category
    category_stats = Tender.objects.values('category')\
        .annotate(count=Count('id'))\
        .order_by('-count')
    
    # Get ISO statistics by division
    division_stats = ISONumber.objects.values('division__name')\
        .annotate(count=Count('id'))\
        .order_by('-count')
    
    # Get monthly trends (last 12 months)
    twelve_months_ago = timezone.now() - timedelta(days=365)
    monthly_tenders = Tender.objects.filter(created_at__gte=twelve_months_ago)\
        .annotate(month=ExtractMonth('created_at'))\
        .values('month')\
        .annotate(count=Count('id'))\
        .order_by('month')
    
    return render(request, 'reports.html', {
        'total_tenders': total_tenders,
        'total_isos': total_isos,
        'completed_tenders': completed_tenders,
        'department_stats': department_stats,
        'category_stats': category_stats,
        'division_stats': division_stats,
        'monthly_tenders': monthly_tenders,
    })

@login_required
@user_passes_test(has_user_role)
def tender_items_view(request, tender_id):
    tender = get_object_or_404(Tender, id=tender_id)
    items = tender.items.all().prefetch_related('vendorbid_set', 'vendorbid_set__vendor')
    
    if request.method == 'POST':
        form = TenderItemForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            item.tender = tender
            item.save()
            messages.success(request, 'Item added successfully!')
            return redirect('tender-items', tender_id=tender_id)
    else:
        form = TenderItemForm()
    
    return render(request, 'tender_items.html', {
        'tender': tender,
        'items': items,
        'form': form
    })

@login_required
@user_passes_test(has_user_role)
def edit_tender_item_view(request, tender_id, item_id):
    tender = get_object_or_404(Tender, id=tender_id)
    item = get_object_or_404(TenderItem, id=item_id, tender=tender)
    
    if request.method == 'POST':
        form = TenderItemForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            messages.success(request, 'Item updated successfully!')
            return redirect('tender-items', tender_id=tender_id)
    
    return redirect('tender-items', tender_id=tender_id)

@login_required
@user_passes_test(has_user_role)
def vendor_bids_view(request, tender_id, item_id):
    tender = get_object_or_404(Tender, id=tender_id)
    item = get_object_or_404(TenderItem, id=item_id, tender=tender)
    bids = VendorBid.objects.filter(tender_item=item).select_related('vendor')
    vendors = Vendor.objects.all()
    
    if request.method == 'POST':
        form = VendorBidForm(request.POST)
        if form.is_valid():
            bid = form.save(commit=False)
            bid.tender = tender
            bid.tender_item = item
            bid.save()
            
            if bid.is_winner:
                # Set other bids for this item as non-winners
                VendorBid.objects.filter(tender_item=item).exclude(id=bid.id).update(is_winner=False)
            
            messages.success(request, 'Bid recorded successfully!')
            return redirect('vendor-bids', tender_id=tender_id, item_id=item_id)
    else:
        form = VendorBidForm()
    
    return render(request, 'vendor_bids.html', {
        'tender': tender,
        'item': item,
        'bids': bids,
        'vendors': vendors,
        'form': form
    })

@login_required
@user_passes_test(has_user_role)
def edit_vendor_bid_view(request, tender_id, item_id, bid_id):
    tender = get_object_or_404(Tender, id=tender_id)
    item = get_object_or_404(TenderItem, id=item_id, tender=tender)
    bid = get_object_or_404(VendorBid, id=bid_id, tender_item=item)
    
    if request.method == 'POST':
        form = VendorBidForm(request.POST, instance=bid)
        if form.is_valid():
            bid = form.save()
            
            if bid.is_winner:
                # Set other bids for this item as non-winners
                VendorBid.objects.filter(tender_item=item).exclude(id=bid.id).update(is_winner=False)
            
            messages.success(request, 'Bid updated successfully!')
            return redirect('vendor-bids', tender_id=tender_id, item_id=item_id)
    
    return redirect('vendor-bids', tender_id=tender_id, item_id=item_id)

@login_required
@user_passes_test(has_user_role)
def framework_agreements_view(request, tender_id):
    tender = get_object_or_404(Tender, id=tender_id)
    agreements = tender.framework_agreements.all().select_related('vendor')
    
    # Get winning vendors from tender items
    winning_vendors = Vendor.objects.all()
    
    if request.method == 'POST':
        form = FrameworkAgreementForm(request.POST)
        if form.is_valid():
            agreement = form.save(commit=False)
            agreement.tender = tender
            agreement.save()
            messages.success(request, 'Framework Agreement created successfully!')
            return redirect('framework-agreements', tender_id=tender_id)
    else:
        form = FrameworkAgreementForm()
    
    return render(request, 'framework_agreements.html', {
        'tender': tender,
        'agreements': agreements,
        'winning_vendors': winning_vendors,
        'form': form
    })

@login_required
@user_passes_test(has_user_role)
def edit_framework_agreement_view(request, tender_id, agreement_id):
    tender = get_object_or_404(Tender, id=tender_id)
    agreement = get_object_or_404(FrameworkAgreement, id=agreement_id, tender=tender)
    
    if request.method == 'POST':
        form = FrameworkAgreementForm(request.POST, instance=agreement)
        if form.is_valid():
            form.save()
            messages.success(request, 'Framework Agreement updated successfully!')
            return redirect('framework-agreements', tender_id=tender_id)
    
    return redirect('framework-agreements', tender_id=tender_id)

@login_required
@user_passes_test(has_user_role)
def chemical_list(request):
    chemicals = Chemical.objects.all().select_related('tender_item__tender')
    
    # Filter by tender item if provided
    tender_item_id = request.GET.get('tender_item')
    if tender_item_id:
        chemicals = chemicals.filter(tender_item_id=tender_item_id)
    
    # Search functionality
    search_query = request.GET.get('search')
    if search_query:
        chemicals = chemicals.filter(
            Q(chemical_name__icontains=search_query) |
            Q(lot_number__icontains=search_query) |
            Q(formula__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(chemicals, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'tender_items': TenderItem.objects.all(),
        'search_query': search_query,
        'tender_item_id': tender_item_id,
    }
    return render(request, 'tender_app/chemical_list.html', context)

@login_required
@user_passes_test(has_user_role)
def chemical_create(request):
    if request.method == 'POST':
        form = ChemicalForm(request.POST)
        if form.is_valid():
            chemical = form.save()
            messages.success(request, 'Chemical created successfully.')
            return redirect('chemical_detail', pk=chemical.pk)
    else:
        form = ChemicalForm()
    
    return render(request, 'tender_app/chemical_form.html', {'form': form, 'title': 'Create Chemical'})

@login_required
@user_passes_test(has_user_role)
def chemical_detail(request, pk):
    chemical = get_object_or_404(Chemical, pk=pk)
    spec_form = ChemicalSpecificationForm()
    
    if request.method == 'POST':
        spec_form = ChemicalSpecificationForm(request.POST)
        if spec_form.is_valid():
            specification = spec_form.save(commit=False)
            specification.chemical = chemical
            specification.save()
            messages.success(request, 'Specification added successfully.')
            return redirect('chemical_detail', pk=pk)
    
    context = {
        'chemical': chemical,
        'spec_form': spec_form,
        'specifications': chemical.specifications.all(),
    }
    return render(request, 'tender_app/chemical_detail.html', context)

@login_required
@user_passes_test(has_user_role)
def chemical_update(request, pk):
    chemical = get_object_or_404(Chemical, pk=pk)
    if request.method == 'POST':
        form = ChemicalForm(request.POST, instance=chemical)
        if form.is_valid():
            form.save()
            messages.success(request, 'Chemical updated successfully.')
            return redirect('chemical_detail', pk=pk)
    else:
        form = ChemicalForm(instance=chemical)
    
    return render(request, 'tender_app/chemical_form.html', {'form': form, 'title': 'Update Chemical'})

@login_required
@user_passes_test(has_user_role)
def chemical_import(request):
    if request.method == 'POST':
        form = ChemicalImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            tender_item = form.cleaned_data['tender_item']
            
            try:
                # Read Excel file
                df = pd.read_excel(excel_file)
                
                for _, row in df.iterrows():
                    # Create Chemical record
                    chemical = Chemical.objects.create(
                        tender_item=tender_item,
                        chemical_name=row['chemical_name'],
                        lot_number=row['lot_number'],
                        formula=row.get('formula', ''),
                        grade=row['grade'],
                        package_size=row['package_size'],
                        quantity=row['quantity']
                    )
                    
                    # Create specifications if they exist
                    spec_columns = ['molar_mass', 'density', 'purity', 'appearance']
                    for spec_type in spec_columns:
                        if spec_type in row and pd.notna(row[spec_type]):
                            ChemicalSpecification.objects.create(
                                chemical=chemical,
                                spec_type=spec_type.upper(),
                                value=str(row[spec_type]),
                                unit=row.get(f'{spec_type}_unit', '')
                            )
                
                messages.success(request, 'Chemicals imported successfully.')
                return redirect('chemical_list')
                
            except Exception as e:
                messages.error(request, f'Error importing chemicals: {str(e)}')
                
    else:
        form = ChemicalImportForm()
    
    return render(request, 'tender_app/chemical_import.html', {'form': form})

@login_required
@user_passes_test(has_user_role)
def chemical_spec_delete(request, pk):
    spec = get_object_or_404(ChemicalSpecification, pk=pk)
    chemical_pk = spec.chemical.pk
    spec.delete()
    messages.success(request, 'Specification deleted successfully.')
    return redirect('chemical_detail', pk=chemical_pk)


# Task Management Views
@login_required
@user_passes_test(has_user_role)
def task_list(request):
    tasks = Task.objects.filter(user=request.user)
    categories = TaskCategory.objects.filter(user=request.user)
    
    # Filter by status if provided
    status_filter = request.GET.get('status')
    if status_filter:
        tasks = tasks.filter(status=status_filter)
    
    # Filter by priority if provided
    priority_filter = request.GET.get('priority')
    if priority_filter:
        tasks = tasks.filter(priority=priority_filter)
        
    context = {
        'tasks': tasks,
        'categories': categories,
        'task_form': TaskForm(initial={'user': request.user}),
        'task_status_choices': Task.STATUS_CHOICES,
        'task_priority_choices': Task.PRIORITY_CHOICES
    }
    return render(request, 'tender_app/task_list.html', context)


@login_required
@user_passes_test(has_user_role)
def task_detail(request, pk):
    task = get_object_or_404(Task, pk=pk, user=request.user)
    comments = task.comments.all()
    
    if request.method == 'POST':
        comment_form = TaskCommentForm(request.POST)
        if comment_form.is_valid():
            comment = comment_form.save(commit=False)
            comment.task = task
            comment.user = request.user
            comment.save()
            messages.success(request, 'Comment added successfully.')
            return redirect('task_detail', pk=task.pk)
    else:
        comment_form = TaskCommentForm()
    
    context = {
        'task': task,
        'comments': comments,
        'comment_form': comment_form
    }
    return render(request, 'tender_app/task_detail.html', context)


@login_required
@user_passes_test(has_user_role)
def task_create(request):
    if request.method == 'POST':
        form = TaskForm(request.POST)
        if form.is_valid():
            task = form.save(commit=False)
            task.user = request.user
            task.save()
            messages.success(request, 'Task created successfully!')
            return redirect('task_list')
    else:
        form = TaskForm(initial={
            'tender': request.GET.get('tender'),
            'vendor': request.GET.get('vendor')
        })
    
    return render(request, 'tender_app/task_form.html', {
        'form': form,
        'action': 'Create'
    })


@login_required
@user_passes_test(has_user_role)
def task_update(request, pk):
    task = get_object_or_404(Task, pk=pk, user=request.user)
    
    if request.method == 'POST':
        form = TaskForm(request.POST, instance=task)
        if form.is_valid():
            form.save()
            messages.success(request, 'Task updated successfully!')
            return redirect('task_detail', pk=task.pk)
    else:
        form = TaskForm(instance=task)
    
    return render(request, 'tender_app/task_form.html', {
        'form': form,
        'task': task,
        'action': 'Update'
    })


@login_required
@user_passes_test(has_user_role)
def task_delete(request, pk):
    task = get_object_or_404(Task, pk=pk, user=request.user)
    
    if request.method == 'POST':
        task.delete()
        messages.success(request, 'Task deleted successfully!')
        return redirect('task_list')
    
    return render(request, 'tender_app/task_confirm_delete.html', {'task': task})


@login_required
@user_passes_test(has_user_role)
def task_status_update(request, pk):
    if request.method == 'POST':
        task = get_object_or_404(Task, pk=pk, user=request.user)
        status = request.POST.get('status')
        
        if status in dict(Task.STATUS_CHOICES).keys():
            task.status = status
            task.save()
            return JsonResponse({'success': True, 'status': task.get_status_display()})
    
    return JsonResponse({'success': False}, status=400)


@login_required
@user_passes_test(has_user_role)
def task_category_create(request):
    if request.method == 'POST':
        form = TaskCategoryForm(request.POST)
        if form.is_valid():
            category = form.save(commit=False)
            category.user = request.user
            category.save()
            messages.success(request, 'Category created successfully!')
            return redirect('task_list')
    else:
        form = TaskCategoryForm()
    
    return render(request, 'tender_app/task_category_form.html', {'form': form})


@login_required
@user_passes_test(has_user_role)
def task_dashboard(request):
    # Get tasks counts by status
    pending_count = Task.objects.filter(user=request.user, status='pending').count()
    in_progress_count = Task.objects.filter(user=request.user, status='in_progress').count()
    completed_count = Task.objects.filter(user=request.user, status='completed').count()
    
    # Get tasks due soon (within 3 days)
    today = datetime.now(timezone.utc)
    due_soon = Task.objects.filter(
        user=request.user,
        due_date__gte=today,
        due_date__lte=today + timedelta(days=3),
        status__in=['pending', 'in_progress']
    )
    
    # Get overdue tasks
    overdue = Task.objects.filter(
        user=request.user,
        due_date__lt=today,
        status__in=['pending', 'in_progress']
    )
    
    context = {
        'pending_count': pending_count,
        'in_progress_count': in_progress_count,
        'completed_count': completed_count,
        'due_soon': due_soon,
        'overdue': overdue,
    }
    
    return render(request, 'tender_app/task_dashboard.html', context)

@login_required
def profile_view(request):
    """View the current user's profile."""
    user_profile = request.user.profile
    return render(request, 'profile.html', {
        'user_profile': user_profile
    })

@login_required
def profile_update_view(request):
    """Update the current user's profile."""
    if request.method == 'POST':
        form = UserProfileForm(request.POST, request.FILES, instance=request.user.profile)
        if form.is_valid():
            form.save()
            messages.success(request, 'Your profile has been updated successfully!')
            return redirect('profile')
    else:
        form = UserProfileForm(instance=request.user.profile)
    
    return render(request, 'profile_update.html', {
        'form': form
    })

@login_required
@user_passes_test(has_user_role)
def vendor_list_view(request):
    """Display a list of all vendors."""
    vendors = Vendor.objects.all().order_by('name')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        vendors = vendors.filter(
            Q(name__icontains=search_query) |
            Q(contact_person__icontains=search_query) |
            Q(email__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(vendors, 10)  # Show 10 vendors per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'vendor_list.html', {
        'page_obj': page_obj,
        'search_query': search_query,
        'total_vendors': vendors.count(),
    })

@login_required
@user_passes_test(has_user_role)
def vendor_detail_view(request, vendor_id):
    """Display details of a specific vendor."""
    vendor = get_object_or_404(Vendor, id=vendor_id)
    
    # Get related tenders through framework agreements
    framework_agreements = vendor.framework_agreements.all().select_related('tender')
    
    # Get related tenders through winning bids
    winning_bids = VendorBid.objects.filter(vendor=vendor, is_winner=True).select_related('tender', 'tender_item').distinct()
    
    # Get associated ISO numbers through tenders
    iso_numbers = []
    for agreement in framework_agreements:
        iso_numbers.extend(list(ISONumber.objects.filter(tender=agreement.tender)))
    
    for bid in winning_bids:
        if bid.tender:
            iso_numbers.extend(list(ISONumber.objects.filter(tender=bid.tender)))
    
    # Remove duplicates
    iso_numbers = list(set(iso_numbers))
    
    return render(request, 'vendor_detail.html', {
        'vendor': vendor,
        'framework_agreements': framework_agreements,
        'winning_bids': winning_bids,
        'iso_numbers': iso_numbers,
    })

@login_required
@user_passes_test(has_user_role)
def vendor_create_view(request):
    """Create a new vendor."""
    if request.method == 'POST':
        form = VendorForm(request.POST)
        if form.is_valid():
            vendor = form.save()
            messages.success(request, f'Vendor "{vendor.name}" created successfully!')
            return redirect('vendor-detail', vendor_id=vendor.id)
    else:
        form = VendorForm()
    
    return render(request, 'vendor_form.html', {
        'form': form,
        'title': 'Add New Vendor',
        'button_text': 'Create Vendor',
    })

@login_required
@user_passes_test(has_user_role)
def vendor_update_view(request, vendor_id):
    """Update an existing vendor."""
    vendor = get_object_or_404(Vendor, id=vendor_id)
    
    if request.method == 'POST':
        form = VendorForm(request.POST, instance=vendor)
        if form.is_valid():
            vendor = form.save()
            messages.success(request, f'Vendor "{vendor.name}" updated successfully!')
            return redirect('vendor-detail', vendor_id=vendor.id)
    else:
        form = VendorForm(instance=vendor)
    
    return render(request, 'vendor_form.html', {
        'form': form,
        'vendor': vendor,
        'title': f'Edit Vendor: {vendor.name}',
        'button_text': 'Update Vendor',
    })

@login_required
@user_passes_test(has_user_role)
def vendor_delete_view(request, vendor_id):
    """Delete an existing vendor."""
    vendor = get_object_or_404(Vendor, id=vendor_id)
    
    if request.method == 'POST':
        vendor_name = vendor.name
        vendor.delete()
        messages.success(request, f'Vendor "{vendor_name}" deleted successfully!')
        return redirect('vendor-list')
    
    return render(request, 'vendor_confirm_delete.html', {
        'vendor': vendor,
    })

def all_framework_agreements_view(request):
    """View to list all framework agreements across all tenders."""
    from datetime import datetime, timedelta
    
    # Get filter parameters first
    today = datetime.now().date()
    soon = today + timedelta(days=30)
    
    # Make efficient DB queries
    agreements = FrameworkAgreement.objects.all().select_related('vendor', 'tender')
    expired = agreements.filter(end_date__lt=today)
    expiring_soon = agreements.filter(end_date__gte=today, end_date__lte=soon)
    active_agreements = agreements.filter(end_date__gt=soon)
    
    grouped_agreements = {
        'all': agreements,
        'active': active_agreements,
        'expiring_soon': expiring_soon,
        'expired': expired,
    }
    
    if request.method == 'POST':
        form = FrameworkAgreementForm(request.POST)
        if form.is_valid():
            agreement = form.save(commit=False)
            tender_id = request.POST.get('tender')
            if tender_id and tender_id.strip():
                try:
                    tender = get_object_or_404(Tender, id=tender_id)
                    agreement.tender = tender
                    agreement.save()
                    messages.success(request, 'Framework Agreement created successfully!')
                    return redirect('all-framework-agreements')
                except Tender.DoesNotExist:
                    messages.error(request, 'Tender not found.')
                except IntegrityError:
                    messages.error(request, 'Database integrity error. This agreement may already exist.')
                except ValidationError as e:
                    messages.error(request, f'Validation error: {str(e)}')
                except Exception as e:
                    messages.error(request, f'Unexpected error: {str(e)}')
                    # Log the unexpected exception
                    logger.error(f"Unexpected error creating agreement: {str(e)}")
            else:
                messages.error(request, 'Please select a tender')
    else:
        form = FrameworkAgreementForm()
    
    return render(request, 'all_framework_agreements.html', {
        'agreements': grouped_agreements,
        'tenders': tenders,
        'form': form,
    })

def generic_list_view(request, model_class, template_name, context_name, 
                      search_fields=None, order_by='-created_at', paginate_by=10):
    objects = model_class.objects.all()
    
    # Apply search if provided
    search_query = request.GET.get('search', '')
    if search_query and search_fields:
        query = Q()
        for field in search_fields:
            query |= Q(**{f"{field}__icontains": search_query})
        objects = objects.filter(query)
    
    # Apply ordering
    objects = objects.order_by(order_by)
    
    # Pagination
    paginator = Paginator(objects, paginate_by)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    
    context = {
        context_name: page_obj,
        'search_query': search_query,
        f'total_{context_name}': objects.count(),
    }
    
    return render(request, template_name, context)

class VendorListView(ListView):
    model = Vendor
    template_name = 'vendor_list.html'
    context_object_name = 'vendors'
    paginate_by = 10
    ordering = ['name']
    
    def get_queryset(self):
        queryset = super().get_queryset()
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(name__icontains=search_query) |
                Q(contact_person__icontains=search_query) |
                Q(email__icontains=search_query)
            )
        return queryset
