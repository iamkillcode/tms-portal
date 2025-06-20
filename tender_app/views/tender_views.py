"""
Tender-related views for the tender application.
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse, HttpResponse, HttpResponseNotAllowed
from django.views.decorators.http import require_POST, require_http_methods
from django.core.paginator import Paginator
from django.db.models import Q
from django.views.generic import DetailView
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from ..models import (
    Tender, Department, Category, TenderTracker, UserProfile, 
    TenderItem, Vendor
)
from ..forms import TenderItemForm


# Helper functions for permission checks
def has_admin_role(user):
    if user.is_superuser:
        return True
    allowed_groups = {'Admin', 'Head of Department', 'Head of Unit'}
    user_groups = set(user.groups.values_list('name', flat=True))
    return bool(allowed_groups & user_groups)


def has_user_role(user):
    if has_admin_role(user):
        return True
    allowed_groups = {'Team Lead', 'Officer'}
    user_groups = set(user.groups.values_list('name', flat=True))
    return bool(allowed_groups & user_groups)


# Admin-only view example
@user_passes_test(has_admin_role)
def admin_dashboard_view(request):
    """Dashboard view for admin users."""
    # Implementation to be filled
    pass


# User-level view example
@user_passes_test(has_user_role)
def user_dashboard_view(request):
    """Dashboard view for regular users."""
    # Implementation to be filled
    pass


# Tender number generator view
@login_required
@user_passes_test(has_user_role)
def tender_generator_view(request):
    """View to generate a new tender number and create a tender."""
    # Ensure user has a profile
    if not hasattr(request.user, 'profile'):
        UserProfile.objects.create(
            user=request.user,
            full_name=f"{request.user.first_name} {request.user.last_name}".strip() or request.user.username
        )
    
    tender_number = None
    error = None
    tender_description = None  # Variable to hold the description

    if request.method == "POST":
        try:
            # Get form inputs
            department_code = request.POST.get("department_code")
            category_code = request.POST.get("category_code")
            procurement_type = request.POST.get("procurement_type")
            tender_description = request.POST.get("tender_description")  # Get the description field
            lot_number = request.POST.get("lot_number")
            amendment_number = request.POST.get("amendment_number")
            call_off_number = request.POST.get("call_off_number")

            # Check if inputs are missing or invalid
            if not department_code or not category_code or not procurement_type:
                raise ValueError("All fields must be provided.")

            prefix = "FDA"
            year = datetime.now().year

            # Retrieve or create a tracker for the current year
            tracker, created = TenderTracker.objects.get_or_create(
                year=year,
                defaults={'last_sequence': 0}  # Initialize sequence if new
            )

            # Increment the global sequence
            tracker.last_sequence += 1
            tracker.save()

            # Generate the tender number
            sequential_number = f"{tracker.last_sequence:04}"  # Zero-padded to 4 digits
            lot_suffix = f" ({int(lot_number):02d})" if lot_number else ""
            amendment_suffix = f" (A{amendment_number})" if amendment_number else ""
            call_off_suffix = f" (CO{call_off_number})" if call_off_number else ""
            tender_number = f"{prefix}/{department_code}/{year}/{category_code}/{procurement_type}-{sequential_number}{lot_suffix}{amendment_suffix}{call_off_suffix}"

            # Save Tender into the database
            department = Department.objects.get(code=department_code)
            category = Category.objects.get(code=category_code)

            tender = Tender.objects.create(
                tender_number=tender_number,
                description=tender_description,
                department=department,
                category=category,
                user=request.user  # Automatically assign logged-in user
            )
            messages.success(request, f'Tender {tender_number} created successfully!')

        except Department.DoesNotExist:
            messages.error(request, 'Invalid department code selected.')
            error = "Invalid department code"
        except Category.DoesNotExist:
            messages.error(request, 'Invalid category code selected.')
            error = "Invalid category code"
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
            error = str(e)

    # Fetch department codes for the dropdown
    department_codes = Department.objects.values_list('code', flat=True)
    categories = Category.objects.all()
    category_data = [{"id": category.code, "text": category.name} for category in categories]

    # Fetch tenders, ordered by most recent
    tenders = Tender.objects.all().order_by('-created_at')

    # Add pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(tenders, 10)  # Show 10 tenders per page
    try:
        tenders = paginator.page(page)
    except:
        tenders = paginator.page(1)

    return render(request, "tender_generator.html", {
        "tender_number": tender_number,
        "error": error,
        "department_codes": department_codes,
        "tender_description": tender_description,  # Pass the description to the template
        "category_data": category_data,
        "tenders": tenders,  # Pass tenders to the template
        "is_paginated": True if tenders.has_other_pages() else False,
    })


@login_required
@user_passes_test(has_user_role)
def tender_activity_view(request):
    """View to display tender activity."""
    search_query = request.GET.get('search', '')
    
    tenders = Tender.objects.all()
    if search_query:
        tenders = tenders.filter(
            Q(tender_number__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(department__name__icontains=search_query)
        )
    
    tenders = tenders.order_by('-created_at')
    
    # Add pagination here too
    paginator = Paginator(tenders, 10)
    page = request.GET.get('page')
    tenders = paginator.get_page(page)
    
    return render(request, 'tender_activity.html', {
        "tenders": tenders,
        "search_query": search_query
    })


@login_required
@user_passes_test(has_user_role)
def tender_list_view(request):
    """View to list all tenders."""
    from ..htmx_utils import htmx_template
    
    @htmx_template('tender_list.html', 'htmx/tender_list_rows.html')
    def _tender_list(request):
        search_query = request.GET.get('search', '')
        
        tenders = Tender.objects.all()
        if search_query:
            tenders = tenders.filter(
                Q(tender_number__icontains=search_query) |
                Q(description__icontains=search_query) |
                Q(department__name__icontains=search_query)
            )
        
        tenders = tenders.order_by('-created_at')
        
        # Get all vendors for the dropdown
        vendors = Vendor.objects.all().order_by('name')
        
        paginator = Paginator(tenders, 10)
        page = request.GET.get('page')
        tenders_page = paginator.get_page(page)
        
        return {
            'tenders': tenders_page,
            'vendors': vendors,
            'search_query': search_query
        }
    
    return _tender_list(request)


@login_required
@user_passes_test(has_user_role)
def tender_update_view(request, tender_id):
    """View to update a tender."""
    tender = get_object_or_404(Tender, id=tender_id)
    
    if request.method == 'POST':
        # Update tender with form data
        tender.category = request.POST.get('category')
        tender.status = request.POST.get('status')
        tender.invitation_date = request.POST.get('invitation_date') or None
        tender.closing_date = request.POST.get('closing_date') or None
        tender.evaluation_date = request.POST.get('evaluation_date') or None
        tender.contract_date = request.POST.get('contract_date') or None
        tender.currency = request.POST.get('currency')
        tender.contract_amount = request.POST.get('contract_amount') or None
        
        # Keep vendor_name field for backward compatibility
        vendor_name = request.POST.get('vendor_name')
        tender.vendor_name = vendor_name
        
        # Create relationship with vendor using proper ForeignKey
        vendor_id = request.POST.get('vendor')
        if vendor_id:
            tender.vendor = get_object_or_404(Vendor, id=vendor_id)
        else:
            tender.vendor = None
        
        tender.save()
        messages.success(request, 'Tender details updated successfully.')
        
    return redirect('tender-list')


@login_required
@user_passes_test(has_admin_role)
def export_tenders_view(request):
    """View to export tenders to Excel."""
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Tender List"

    # Define headers and styles
    headers = [
        'Tender Number',
        'Description',
        'Department',
        'Category',
        'Status',
        'Officer',
        'Invitation Date',
        'Closing Date',
        'Evaluation Date',
        'Contract Date',
        'Currency',
        'Contract Amount',
        'Vendor/Consultant',
        'PO Number',
        'PO Date',
        'SRA Number',
        'SRA Date',
        'Payment Amount',
        'File Name',
        'File Number',
        'Created Date'
    ]

    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Get all tenders
    tenders = Tender.objects.all().order_by('-created_at')

    # Write data with timezone handling
    for row, tender in enumerate(tenders, 2):
        ws.cell(row=row, column=1, value=tender.tender_number)
        ws.cell(row=row, column=2, value=tender.description)
        ws.cell(row=row, column=3, value=tender.department.name)
        ws.cell(row=row, column=4, value=tender.category)
        ws.cell(row=row, column=5, value=tender.status)
        
        try:
            officer_name = tender.user.profile.full_name
        except:
            officer_name = tender.user.get_full_name() or tender.user.username
        
        ws.cell(row=row, column=6, value=officer_name)
        
        # Handle dates with timezone conversion
        date_fields = [
            tender.invitation_date,
            tender.closing_date,
            tender.evaluation_date,
            tender.contract_date,
            tender.po_date,
            tender.sra_date,
            tender.payment_memo_date
        ]
        
        for col, date in enumerate(date_fields, 7):
            if date:
                # Date fields don't need timezone handling as they're already timezone-naive
                ws.cell(row=row, column=col, value=date)

        # Handle numeric and text fields
        ws.cell(row=row, column=11, value=tender.currency)
        ws.cell(row=row, column=12, value=tender.contract_amount)
        ws.cell(row=row, column=13, value=tender.vendor_name)
        ws.cell(row=row, column=14, value=tender.po_number)
        ws.cell(row=row, column=16, value=tender.sra_number)
        ws.cell(row=row, column=18, value=tender.payment_amount)
        ws.cell(row=row, column=19, value=tender.file_name)
        ws.cell(row=row, column=20, value=tender.file_number)

        # Handle created_at datetime with timezone conversion
        created_at = tender.created_at
        if hasattr(created_at, 'tzinfo') and created_at.tzinfo:
            created_at = created_at.astimezone(datetime.timezone.utc).replace(tzinfo=None)

        ws.cell(row=row, column=21, value=created_at)

        # Format the date cell
        date_cell = ws.cell(row=row, column=21)
        date_cell.number_format = 'YYYY-MM-DD HH:MM:SS'

    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=tender_list_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    try:
        wb.save(response)
    except Exception as e:
        messages.error(request, f"Error exporting tenders: {str(e)}")
        return redirect('tender-list')

    return response


@login_required
@user_passes_test(has_user_role)
def tender_items_view(request, tender_id):
    """View to list and manage items for a specific tender."""
    from ..htmx_utils import htmx_template, is_htmx_request
    
    tender = get_object_or_404(Tender, id=tender_id)
    items = tender.items.all().prefetch_related('vendorbid_set', 'vendorbid_set__vendor')
    
    @htmx_template('tender_items.html', 'htmx/tender_items_form.html')
    def process_form(request):
        if request.method == 'POST':
            form = TenderItemForm(request.POST)
            if form.is_valid():
                item = form.save(commit=False)
                item.tender = tender
                item.save()
                messages.success(request, 'Item added successfully!')
                
                # If HTMX request, return the updated list
                if is_htmx_request(request):
                    response = render(request, 'htmx/tender_items_list.html', {
                        'tender': tender, 
                        'items': items
                    })
                    response['HX-Trigger'] = '{"itemAdded": true}'
                    return response
                    
                return redirect('tender-items', tender_id=tender_id)
        else:
            form = TenderItemForm()
            return {
                'tender': tender,
                'items': items,
                'form': form
            }
                
    return process_form(request)


@login_required
@user_passes_test(has_user_role)
def tender_item_edit_view(request, tender_id, item_id):
    """View to edit a tender item with HTMX support."""
    from ..htmx_utils import is_htmx_request
    
    tender = get_object_or_404(Tender, id=tender_id)
    item = get_object_or_404(TenderItem, id=item_id, tender=tender)
    
    if request.method == 'GET':
        form = TenderItemForm(instance=item)
        return render(request, 'htmx/tender_items_edit.html', {
            'tender': tender,
            'item': item,
            'form': form
        })
    
    elif request.method == 'PUT':
        # Use request.PUT which is set by our middleware
        form = TenderItemForm(request.PUT, instance=item)
        if form.is_valid():
            form.save()
            messages.success(request, 'Item updated successfully!')
            
            # Return the updated row for HTMX swap
            items = [item]  # Just this one item for the response
            return render(request, 'htmx/tender_items_list.html', {
                'tender': tender,
                'items': items
            })
        else:
            # Form has errors
            return render(request, 'htmx/tender_items_edit.html', {
                'tender': tender,
                'item': item,
                'form': form
            })
    
    # Fall back for unsupported methods
    return HttpResponseNotAllowed(['GET', 'PUT'])


@login_required
@user_passes_test(has_user_role)
def tender_item_delete_view(request, tender_id, item_id):
    """View to delete a tender item with HTMX support."""
    tender = get_object_or_404(Tender, id=tender_id)
    item = get_object_or_404(TenderItem, id=item_id, tender=tender)
    
    if request.method == 'DELETE':
        item.delete()
        messages.success(request, 'Item deleted successfully!')
        
        # Return an empty response for HTMX to remove the element
        return HttpResponse(status=200)
    
    return HttpResponseNotAllowed(['DELETE'])
    
class TenderDetailView(DetailView):
    """Detail view for a tender."""
    model = Tender
    template_name = 'tender_app/tender_detail.html'
    context_object_name = 'tender'
    
@login_required
@user_passes_test(has_user_role)
def search_view(request):
    """View to search across tenders and ISOs with HTMX support."""
    from ..htmx_utils import is_htmx_request
    
    search_query = request.GET.get('search', '')
    
    tenders = []
    isos = []
    
    if search_query:
        from ..models import ISONumber
        
        tenders = Tender.objects.filter(
            Q(tender_number__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(department__name__icontains=search_query) |
            Q(category__icontains=search_query)
        ).order_by('-created_at')
        
        isos = ISONumber.objects.filter(
            Q(iso_number__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(division__name__icontains=search_query) |
            Q(department__name__icontains=search_query)
        ).order_by('-date_created')
    
    # Check if it's an HTMX request and which target is specified
    if is_htmx_request(request):
        target = request.headers.get('HX-Target')
        
        if target == 'tender-results':
            return render(request, 'htmx/search_tenders_results.html', {
                'tenders': tenders,
            })
        elif target == 'iso-results':
            return render(request, 'htmx/search_iso_results.html', {
                'isos': isos,
            })
      # Regular request, render the full template
    return render(request, 'search.html', {
        'search_query': search_query,
        'tenders': tenders,
        'isos': isos
    })
