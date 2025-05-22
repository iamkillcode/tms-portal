from django.shortcuts import render
from django.contrib.auth.decorators import user_passes_test, login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.db.models import Q
from datetime import datetime, timedelta
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from .models import AuditLog

@staff_member_required
def audit_trail_dashboard(request):
    """
    Dashboard view for audit logs with filtering options.
    """
    # Get filters from request
    user_id = request.GET.get('user_id')
    action_type = request.GET.get('action_type')
    app_name = request.GET.get('app_name')
    model_name = request.GET.get('model_name')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    search_query = request.GET.get('search', '')
    
    # Base queryset
    logs = AuditLog.objects.all()
    
    # Apply filters
    if user_id:
        logs = logs.filter(user_id=user_id)
    
    if action_type:
        logs = logs.filter(action_type=action_type)
    
    if app_name:
        logs = logs.filter(app_name=app_name)
    
    if model_name:
        logs = logs.filter(model_name=model_name)
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            logs = logs.filter(timestamp__gte=date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            # Add one day to include the entire end date
            date_to_obj = date_to_obj + timedelta(days=1)
            logs = logs.filter(timestamp__lt=date_to_obj)
        except ValueError:
            pass
    
    if search_query:
        logs = logs.filter(
            Q(user__username__icontains=search_query) |
            Q(action_detail__icontains=search_query) |
            Q(object_repr__icontains=search_query) |
            Q(ip_address__icontains=search_query)
        )
    
    # Order by timestamp (most recent first)
    logs = logs.order_by('-timestamp')
    
    # Pagination
    paginator = Paginator(logs, 50)  # Show 50 logs per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get unique values for filters
    unique_users = AuditLog.objects.values_list('user__username', 'user_id').distinct().order_by('user__username')
    unique_action_types = AuditLog.objects.values_list('action_type', flat=True).distinct().order_by('action_type')
    unique_app_names = AuditLog.objects.values_list('app_name', flat=True).distinct().order_by('app_name')
    unique_model_names = AuditLog.objects.values_list('model_name', flat=True).distinct().order_by('model_name')
    
    context = {
        'page_obj': page_obj,
        'unique_users': unique_users,
        'unique_action_types': unique_action_types,
        'unique_app_names': unique_app_names,
        'unique_model_names': unique_model_names,
        'selected_user_id': user_id,
        'selected_action_type': action_type,
        'selected_app_name': app_name,
        'selected_model_name': model_name,
        'date_from': date_from,
        'date_to': date_to,
        'search_query': search_query,
    }
    
    return render(request, 'audit_trail/dashboard.html', context)

@staff_member_required
def export_audit_logs(request, format='csv'):
    """
    Export audit logs to CSV or Excel.
    """
    # Get filters from request (same as dashboard)
    user_id = request.GET.get('user_id')
    action_type = request.GET.get('action_type')
    app_name = request.GET.get('app_name')
    model_name = request.GET.get('model_name')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    search_query = request.GET.get('search', '')
    
    # Base queryset
    logs = AuditLog.objects.all()
    
    # Apply filters (same as dashboard)
    if user_id:
        logs = logs.filter(user_id=user_id)
    
    if action_type:
        logs = logs.filter(action_type=action_type)
    
    if app_name:
        logs = logs.filter(app_name=app_name)
    
    if model_name:
        logs = logs.filter(model_name=model_name)
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            logs = logs.filter(timestamp__gte=date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            date_to_obj = date_to_obj + timedelta(days=1)
            logs = logs.filter(timestamp__lt=date_to_obj)
        except ValueError:
            pass
    
    if search_query:
        logs = logs.filter(
            Q(user__username__icontains=search_query) |
            Q(action_detail__icontains=search_query) |
            Q(object_repr__icontains=search_query) |
            Q(ip_address__icontains=search_query)
        )
    
    # Order by timestamp
    logs = logs.order_by('-timestamp')
    
    # Define column headers
    headers = [
        'Timestamp',
        'User',
        'Action Type',
        'App',
        'Model',
        'Object',
        'Action Detail',
        'IP Address',
        'User Agent'
    ]
    
    if format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename=audit_logs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        writer = csv.writer(response)
        
        # Write header
        writer.writerow(headers)
        
        # Write data rows
        for log in logs:
            user_name = log.user.username if log.user else 'Anonymous'
            writer.writerow([
                log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                user_name,
                log.get_action_type_display(),
                log.app_name,
                log.model_name,
                log.object_repr,
                log.action_detail,
                log.ip_address,
                log.user_agent
            ])
        
        return response
        
    elif format == 'excel':
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Audit Logs"
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Write data rows
        for row_num, log in enumerate(logs, 2):
            user_name = log.user.username if log.user else 'Anonymous'
            
            ws.cell(row=row_num, column=1, value=log.timestamp)
            ws.cell(row=row_num, column=2, value=user_name)
            ws.cell(row=row_num, column=3, value=log.get_action_type_display())
            ws.cell(row=row_num, column=4, value=log.app_name)
            ws.cell(row=row_num, column=5, value=log.model_name)
            ws.cell(row=row_num, column=6, value=log.object_repr)
            ws.cell(row=row_num, column=7, value=log.action_detail)
            ws.cell(row=row_num, column=8, value=log.ip_address)
            ws.cell(row=row_num, column=9, value=log.user_agent)
        
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = openpyxl.utils.get_column_letter(col)
            ws.column_dimensions[column_letter].width = 20
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=audit_logs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        wb.save(response)
        return response
    
    # Default fallback
    return HttpResponse("Invalid export format requested", status=400)
